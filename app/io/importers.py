"""Parse and validate vocabulary import files (CSV, TSV, Excel, JSON)."""
from __future__ import annotations

import csv
import io
import json
import tempfile
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from ..extensions import db
from ..models.vocabulary import VocabularyEntry

REQUIRED_FIELDS = {"word", "meaning", "target_language"}
OPTIONAL_FIELDS = {"lecture", "date_added", "synonyms", "antonyms",
                   "translation_en", "metadata_usage"}
ALL_FIELDS = REQUIRED_FIELDS | OPTIONAL_FIELDS


def _normalise_row(raw: dict[str, Any]) -> dict[str, Any]:
    """Strip whitespace from string values and lowercase field names."""
    return {k.strip().lower(): (v.strip() if isinstance(v, str) else v)
            for k, v in raw.items()}


def _parse_synonyms(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    if isinstance(value, str) and value.strip():
        return [s.strip() for s in value.split("|") if s.strip()]
    return []


def _parse_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value), fmt).date()
        except ValueError:
            pass
    return None


# ---------------------------------------------------------------------------
# Format-specific parsers — each returns list[dict]
# ---------------------------------------------------------------------------

def _parse_csv_bytes(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig")  # strip BOM if present
    reader = csv.DictReader(io.StringIO(text))
    return [_normalise_row(row) for row in reader]


def _parse_tsv_bytes(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text), delimiter="\t")
    return [_normalise_row(row) for row in reader]


def _parse_excel_bytes(data: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    rows: list[dict] = []
    for ws in wb.worksheets:
        header: list[str] | None = None
        for raw_row in ws.iter_rows(values_only=True):
            values = [str(c).strip() if c is not None else "" for c in raw_row]
            if header is None:
                header = [v.lower() for v in values]
                continue
            row = dict(zip(header, values))
            rows.append(_normalise_row(row))
    wb.close()
    return rows


def _parse_json_bytes(data: bytes) -> list[dict]:
    payload = json.loads(data.decode("utf-8"))
    if not isinstance(payload, list):
        raise ValueError("JSON root must be an array of objects.")
    return [_normalise_row(item) for item in payload]


def parse_file(filename: str, data: bytes) -> tuple[list[dict], list[str]]:
    """Detect format from filename extension and parse bytes into row dicts.

    Returns (rows, global_errors).  global_errors is non-empty only when the
    file cannot be parsed at all.
    """
    ext = Path(filename).suffix.lower()
    try:
        if ext == ".csv":
            return _parse_csv_bytes(data), []
        if ext in (".tsv", ".tab"):
            return _parse_tsv_bytes(data), []
        if ext in (".xlsx", ".xlsm"):
            return _parse_excel_bytes(data), []
        if ext == ".json":
            return _parse_json_bytes(data), []
        return [], [f"Unsupported file type: {ext!r}"]
    except Exception as exc:  # noqa: BLE001
        return [], [f"Could not parse file: {exc}"]


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_rows(
    rows: list[dict],
    user_id: str,
) -> tuple[list[dict], list[dict], list[dict]]:
    """Classify rows into valid, duplicate, and error buckets.

    Returns:
        valid_rows    — list of cleaned row dicts ready for import
        duplicate_rows — rows whose word+lecture already exist for this user
        error_rows     — rows with missing required fields or other issues
    """
    # Build a set of existing (word, lecture) for this user
    existing: set[tuple[str, str]] = set()
    for e in VocabularyEntry.query.filter_by(user_id=user_id).with_entities(
        VocabularyEntry.word, VocabularyEntry.lecture
    ):
        existing.add((e.word.lower(), e.lecture.lower()))

    valid: list[dict] = []
    duplicates: list[dict] = []
    errors: list[dict] = []

    for i, row in enumerate(rows):
        missing = [f for f in REQUIRED_FIELDS if not row.get(f)]
        if missing:
            errors.append({**row, "_row": i + 1, "_error": f"Missing: {', '.join(missing)}"})
            continue

        lecture = row.get("lecture", "General").strip() or "General"
        word = row["word"].strip()
        if (word.lower(), lecture.lower()) in existing:
            duplicates.append({**row, "_row": i + 1, "lecture": lecture})
            continue

        valid.append({**row, "lecture": lecture, "word": word})

    return valid, duplicates, errors


# ---------------------------------------------------------------------------
# Commit
# ---------------------------------------------------------------------------

def commit_import(
    rows: list[dict],
    user_id: str,
    overwrite: bool = False,
) -> int:
    """Insert rows into the DB.  If overwrite=True, update duplicate entries
    in place rather than skipping them.  Returns number of rows inserted/updated.
    """
    count = 0
    today = date.today()

    existing_map: dict[tuple[str, str], VocabularyEntry] = {}
    if overwrite:
        for e in VocabularyEntry.query.filter_by(user_id=user_id):
            existing_map[(e.word.lower(), e.lecture.lower())] = e

    for row in rows:
        word = row["word"].strip()
        lecture = (row.get("lecture") or "General").strip()
        synonyms = _parse_synonyms(row.get("synonyms", ""))
        antonyms = _parse_synonyms(row.get("antonyms", ""))
        date_added = _parse_date(row.get("date_added")) or today

        key = (word.lower(), lecture.lower())
        if overwrite and key in existing_map:
            e = existing_map[key]
            e.meaning = row["meaning"].strip()
            e.translation_en = (row.get("translation_en") or "").strip()
            e.metadata_usage = (row.get("metadata_usage") or "").strip() or None
            e.synonyms = synonyms
            e.antonyms = antonyms
            e.target_language = row["target_language"].strip()
            e.date_added = date_added
            e.updated_at = datetime.now(timezone.utc)
        else:
            e = VocabularyEntry(
                id=str(uuid.uuid4()),
                user_id=user_id,
                word=word,
                lecture=lecture,
                date_added=date_added,
                meaning=row["meaning"].strip(),
                translation_en=(row.get("translation_en") or "").strip(),
                metadata_usage=(row.get("metadata_usage") or "").strip() or None,
                synonyms=synonyms,
                antonyms=antonyms,
                target_language=row["target_language"].strip(),
            )
            db.session.add(e)
        count += 1

    db.session.commit()
    return count


# ---------------------------------------------------------------------------
# Temp-file helpers (avoid large session cookies)
# ---------------------------------------------------------------------------

def save_pending(rows: list[dict]) -> str:
    """Serialise rows to a temp file; return the file path."""
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".json", prefix="vocabimport_", delete=False, encoding="utf-8"
    ) as f:
        json.dump(rows, f, ensure_ascii=False, default=str)
        return f.name


def load_pending(path: str) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def delete_pending(path: str) -> None:
    try:
        Path(path).unlink(missing_ok=True)
    except OSError:
        pass
