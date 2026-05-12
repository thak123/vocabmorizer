"""Export vocabulary entries to CSV, TSV, Excel, or JSON."""
from __future__ import annotations

import csv
import io
import json

from ..models.vocabulary import VocabularyEntry

# Column order matches the spec table header
EXPORT_COLUMNS = [
    "lecture", "date_added", "word", "synonyms", "antonyms",
    "meaning", "translation_en", "metadata_usage", "target_language",
]


def _row(entry: VocabularyEntry) -> dict:
    return {
        "lecture": entry.lecture,
        "date_added": str(entry.date_added),
        "word": entry.word,
        "synonyms": "|".join(entry.synonyms or []),
        "antonyms": "|".join(entry.antonyms or []),
        "meaning": entry.meaning,
        "translation_en": entry.translation_en,
        "metadata_usage": entry.metadata_usage or "",
        "target_language": entry.target_language,
    }


def export_csv(entries: list[VocabularyEntry]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()
    for e in entries:
        writer.writerow(_row(e))
    # UTF-8 BOM so Excel opens without mangling special chars
    return ("﻿" + buf.getvalue()).encode("utf-8")


def export_tsv(entries: list[VocabularyEntry]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS, delimiter="\t")
    writer.writeheader()
    for e in entries:
        writer.writerow(_row(e))
    return buf.getvalue().encode("utf-8")


def export_excel(entries: list[VocabularyEntry]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Group by lecture
    by_lecture: dict[str, list[VocabularyEntry]] = {}
    for e in entries:
        by_lecture.setdefault(e.lecture, []).append(e)

    if not by_lecture:
        by_lecture["Sheet1"] = []

    for lecture, lec_entries in by_lecture.items():
        ws = wb.create_sheet(title=lecture[:31])  # sheet names max 31 chars
        ws.append(EXPORT_COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for e in lec_entries:
            r = _row(e)
            ws.append([r[col] for col in EXPORT_COLUMNS])
        # Auto-width (approximate)
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_json(entries: list[VocabularyEntry]) -> bytes:
    out = []
    for e in entries:
        obj = _row(e)
        obj["id"] = e.id
        obj["synonyms"] = e.synonyms or []
        obj["antonyms"] = e.antonyms or []
        if e.review_stats:
            s = e.review_stats
            obj["review_stats"] = {
                "times_reviewed": s.times_reviewed,
                "correct_count": s.correct_count,
                "consecutive_correct": s.consecutive_correct,
                "is_problematic": s.is_problematic,
                "ease_factor": s.ease_factor,
                "interval_days": s.interval_days,
                "next_review_date": str(s.next_review_date) if s.next_review_date else None,
                "last_reviewed": s.last_reviewed.isoformat() if s.last_reviewed else None,
            }
        out.append(obj)
    return json.dumps(out, ensure_ascii=False, indent=2).encode("utf-8")
